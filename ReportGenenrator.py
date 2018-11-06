#  Working envrionment: Python 2.7
import re
from os import listdir,getcwd
from openpyxl import load_workbook
from openpyxl.styles import colors,PatternFill
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdftypes import resolve1
from pdfminer.psparser import PSLiteral, literal_name

STYLE_RED = PatternFill(fill_type='solid',fgColor=colors.RED)
LOGIC_TEXT = r' (and|or)'
VERDICT_TEXT = r' (supported|not supported|present|not present or not active|disabled or not supported)'

# connect to template excel
def template_open():
    global expectTemplate, realTemplate
    try:
        expectTemplate = load_workbook('template_RGpath.xlsx')
        realTemplate = load_workbook('template_realResult.xlsx')
    except IOError as e:
        raw_input(e)
        exit()

# multiple replace
def multi_replace(str, arg1, arg2, *args):
    str = str.replace(arg1, arg2)
    for arg in args:
        str = str.replace(arg, arg2)
    return str

#  extract data from pdf-acroform fields
def load_fields_from_pdf(field, T=''):
    #  Recursively load form fields
    form = field.get('Kids', None)
    t = field.get('T')
    if t is None:
        t = T
    else:
        #  Add its father name
        t = T + '.' + t if T != '' else t
    """ Following is to repeat fields that have "Kids", now is commented because 
    1. There could be multiple fileds who shared the same field name.
    2. For buttons, the parents has "V" value already, don't need to dig in Kids.
    """
    # if form and t:
    #     return [load_fields_from_pdf(resolve1(f), t) for f in form]
    # else:
    # Some field types, like signatures, need extra resolving
    value = resolve1(field.get('AS')) if resolve1(field.get('AS')) is not None else resolve1(field.get('V'))
    #  if output is PSLiteral type, transfer it into str type through "literal_name" function
    if isinstance(value, PSLiteral):
        return (t, literal_name(value))
    else:
        return (t, resolve1(value))

#  split data into dictionary
def split_data(field, d={}):
    flag = True if len(field) != 2 else False
    if flag:
        for f in field:
            split_data(f, d)
    elif isinstance(field[0], (tuple, list)):
        for f in field:
            split_data(f, d)
    else:
        key = field[0] if field[0] is not None else field[0]
        d[key] = field[1]
    return d

#  load ICS data from decrypted pdf docutment
def load_data_from_pdf(pdf):
    with open(pdf, 'rb') as file:
        parser = PDFParser(file)
        doc = PDFDocument(parser)
        parser.set_document(doc)
        outcome = [load_fields_from_pdf(resolve1(f)) for f in resolve1(doc.catalog['AcroForm'])['Fields']]
    # format the outcome of data extract from ics pdf
    outcome = split_data(outcome)
    if outcome['Max Dynamic Reader Limit sets supported']:
        outcome['Max Dynamic Reader Limit sets supported'] = True if int(outcome['Max Dynamic Reader Limit sets supported']) > 4 else False
    if outcome['Product Configuration']:
        outcome['Product Configuration'] = True if outcome['Product Configuration'] == '(A) PCDA (IRWIN Reader) / S-ICR' else False
    for key in outcome:
        if outcome[key] == 'Yes':
            outcome[key] = True
        elif outcome[key] in ['Off', 'No']:
            outcome[key] = False
    return outcome

# Generate results for sub-condition
def ics_static_save(raw):
    temp_sheet = expectTemplate['ICS_Config_Static'] # list all possible configurtion
    for key in temp_sheet['B']:
        if key.value is None:
            verdict = 'True'
        elif re.search(r'.* and .*', key.value):
            verdict = ''
            pieces = key.value.split(' and ')
            for p in pieces:
                verdict += str(raw[p])
                verdict += ' and '
            verdict = eval(verdict[:-5])
        elif re.search(r'.* or .*', key.value):
            verdict = ''
            pieces = key.value.split(' or ')
            for p in pieces:
                verdict += str(raw[p])
                verdict += ' or '
            verdict = eval(verdict[:-4])
        else:
            verdict = raw[key.value] if key.value != "MSD Support" else (not raw[key.value])
            if key.value == 'Max Dynamic Reader Limit sets supported' and temp_sheet['C15'].value == 'False':
                verdict = 'False'
        if verdict is None:
            raise Exception('There is blank to be filled at row %s. Please retry after fill it.'%key.row)
        temp_sheet['C' + str(key.row)] = str(verdict)

def gen_expect_result():
    icsFile = raw_input('Please input the path of decrypted ICS doc.\n')
    icsData = load_data_from_pdf(icsFile)
    ics_static_save(icsData)

    temp_sheet = expectTemplate['ICS_Config_Static'] # list all possible configurtion
    for path_sheet in [expectTemplate['MSD Path(Online Only)'], expectTemplate['MSD Path(Online Capable)'], expectTemplate['qVSDC Path']]:
        for condition_cell in path_sheet['B']:
            condition = condition_cell.value  # get condition description from template
            #outcome = ''  # contain logic method and verdict calculation the final result
            if condition not in [None, 'CONDITIONS ']:
                if (temp_sheet['C2'].value == 'True' and path_sheet in [expectTemplate['MSD Path(Online Only)'], expectTemplate['MSD Path(Online Capable)']])\
                        or (temp_sheet['C3'].value == 'True' and path_sheet == expectTemplate['MSD Path(Online Only)'])\
                        or (temp_sheet['C4'].value == 'True' and path_sheet == expectTemplate['MSD Path(Online Capable)']):
                    condition = False
                else:
                    # format conditions
                    condition = condition.lower()
                    condition = multi_replace(condition, '\n', ' ', '   ', '  ')    # remove line break and make sure there's only one blank between words
                    condition = multi_replace(condition, '[', '', ']', '(', ')')    # remove all parentheses
                    condition = condition.replace('disbaled', 'disabled')         # fix TYPOs
                    # translate conditions
                    for item in temp_sheet['A']:
                        if item.value and re.search(item.value, condition):
                            if re.search(item.value + VERDICT_TEXT, condition):
                                modifier = re.search(item.value + VERDICT_TEXT, condition).group(0)
                                modifier = modifier.replace(item.value, '').strip()
                                pass
                            else:
                                condition = condition.replace(item.value, item.value + ' supported')
                                modifier = 'supported'
                            if re.search(item.value + VERDICT_TEXT + LOGIC_TEXT, condition):
                                pass
                            else:
                                condition = re.sub(item.value + VERDICT_TEXT, r'\g<0> and', condition)
                            if modifier == 'supported' or modifier == 'present':
                                verdict = temp_sheet['C' + str(item.row)].value
                            elif modifier == 'not supported':
                                verdict = True if temp_sheet['C' + str(item.row)].value == 'True' else False
                                verdict = str(not verdict)
                            elif modifier in ['not present or not active', 'disabled or not supported']:
                                condition = re.sub(item.value + VERDICT_TEXT + LOGIC_TEXT, '', condition)
                                continue
                            condition = re.sub(item.value + VERDICT_TEXT, verdict, condition)
                    condition = condition.strip()
                    condition = condition[:-4]
                    condition = True if not condition else eval(condition)
                subcondition_area = [sub.column for sub in path_sheet[condition_cell.row] if path_sheet[str(sub.column) + '2'].value == 'Y or N/A']
                result_area = [res.column for res in path_sheet[condition_cell.row] if path_sheet[str(res.column) + '2'].value == 'RESULT']
                for index, subcondition in enumerate(subcondition_area):
                    subcondition = True if path_sheet[subcondition + str(condition_cell.row)].value == 'Y' else False
                    if path_sheet[subcondition_area[index] + '1'].value == 'qVSDC/MSD active mode' and temp_sheet['C31'].value == 'False':
                        subcondition = False
                    path_sheet[result_area[index] + str(condition_cell.row)].value = 'Pass' if subcondition and condition else 'NA'
    expectTemplate.save('expectResult.xlsx')
    expectTemplate.close()
    realTemplate.close()


def load_results_from_icstool(input, path_sheet, col):
    #results = {}
    col = chr(ord(col) + 1)
    try:
        icc_workbook = load_workbook(input)
        icc_results = icc_workbook['Test Results']
    except IOError as e:
        raw_input(e)
        exit()
    # for tc in icc_results['B']:
    #     if tc.value not in [None, 'TestCase'] and icc_results['C' + str(tc.row)].value != 'N/A' and (tc.value not in results or results[tc.value] != 'Fail'):
    #         results[tc.value] = icc_results['C' + str(tc.row)].value
    for tc_real in path_sheet['A']:
        if tc_real.value not in [None, 'TEST CASES']:
            tc_realv = tc_real.value.replace('.', '')
            for tc_icc in icc_results['A']:
                visaTCname = icc_results['B' + str(tc_icc.row)].value.split('.')[0]
                tc_iccv = tc_icc.value.replace('NO SCRIPT', visaTCname)
                if re.search(tc_realv, tc_iccv) and icc_results['C' + str(tc_icc.row)].value != 'N/A' and path_sheet[col + str(tc_real.row)].value != 'Fail':
                    path_sheet[col + str(tc_real.row)].value = icc_results['C' + str(tc_icc.row)].value
            if not path_sheet[col + str(tc_real.row)].value or path_sheet[col + str(tc_real.row)].value.strip() == '':
                path_sheet[col + str(tc_real.row)].value = 'NA'
    icc_workbook.close()


def gen_real_result():
    Filepath = getcwd()
    #titles = [t.value for t in realTemplate['Titles']['A']]  # list all possible titles
    #f = [f.value for f in realTemplate['Titles']['B']]
    for path_sheet in [realTemplate['MSD Path(Online Only)'], realTemplate['MSD Path(Online Capable)'], realTemplate['qVSDC Path']]:
        for title in path_sheet['1']:
            FilenameCount = [] # initial fine name counter
            if title.value is not None:
                title_index = [f.row for f in realTemplate['Titles']['A'] if (path_sheet.title + title.value) == f.value]
                if len(title_index) == 1:
                    finder = realTemplate['Titles']['B' + str(title_index[0])].value
                    for Filename in listdir(Filepath):
                        if re.search(r'^(?!.~).*' + str(finder), Filename):
                            FilenameCount.append(Filename)
                    if len(FilenameCount) == 0:
                        for tc_real in path_sheet['A']:
                            if tc_real.value not in [None, 'TEST CASES']:
                                path_sheet[chr(ord(title.column) + 1) + str(tc_real.row)].value = 'NA'
                    elif len(FilenameCount) == 1:
                        load_results_from_icstool(FilenameCount[0], path_sheet, title.column)#write according file
                    elif len(FilenameCount) > 1:
                        raise Exception('There are more than one ICC solution excels which share the same name!')
                elif len(title_index) > 1:
                    raise Exception('Shall not appear the same value in "Titles" worksheet!')
    realTemplate.save('realResult.xlsx')
    realTemplate.close()
    expectTemplate.close()


def cmp_results():
    try:
        expectResult = load_workbook('expectResult.xlsx')
        realResult = load_workbook('realResult.xlsx')
    except IOError as e:
        raw_input(e)
        exit()
    for pathsheet_real in [realResult['MSD Path(Online Only)'], realResult['MSD Path(Online Capable)'], realResult['qVSDC Path']]:
        result_area = [res.column for res in pathsheet_real['2'] if res.value == 'RESULT']
        for resultCol in result_area:
            for resultRow in range(3, pathsheet_real.max_row):
                cmpReal = pathsheet_real[str(resultCol) + str(resultRow)].value
                cmpExpect = expectResult[pathsheet_real.title][str(resultCol) + str(resultRow)].value
                if cmpReal != cmpExpect:
                    pathsheet_real[str(resultCol) + str(resultRow)].fill = STYLE_RED
    realResult.save('realResult.xlsx')
    realResult.close()
    expectResult.close()

def user_interface():
    option = raw_input('***  Which service do you prefer  ***\n1. Generate expectResult from decrypted ICS doc.\n'
                       '2. Generate realResult from ICC solution tool.\n'
                       '3. Compare expectResult and realResult(you may need to perform service 1 and 2 first)\n'
                       '4. Perform all above services.\n'
                       '5. Exit\n')
    if option == '1':
        template_open()
        gen_expect_result()
        return user_interface()
    elif option == '2':
        template_open()
        gen_real_result()
        return user_interface()
    elif option == '3':
        cmp_results()
        return user_interface()
    elif option == '4':
        template_open()
        gen_expect_result()
        template_open()
        gen_real_result()
        cmp_results()
        return user_interface()
    elif option == '5':
        exit()
    else:
        print('Please enter correct service index number.')
        return user_interface()
    # ics = load_data_from_pdf('out1.pdf')
    # ics_static_save(ics)
    # gen_expect_result()
    # gen_real_result()
    # cmp_results()

if __name__ == '__main__':
    user_interface()
