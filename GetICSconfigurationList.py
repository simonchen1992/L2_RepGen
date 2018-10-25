#  Working envrionment: Python 2.7
import re
from openpyxl import load_workbook
from openpyxl.styles import colors,PatternFill
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdftypes import resolve1
from pdfminer.psparser import PSLiteral, literal_name
style = PatternFill(fill_type='solid',fgColor=colors.RED)



# multiple replace
def multi_replace(str, arg1, arg2, *args):
    str = str.replace(arg1, arg2)
    for arg in args:
        str = str.replace(arg, arg2)
    return str


#  get all non-repetitive [ICS Configuration] from checklist
def define_ics_configuration():
    list = []
    try:
        checklist = load_workbook('template_RGpath.xlsx')
    except IOError as e:
        raw_input(e)
        exit()
    for checklist_sheet in [checklist['MSD Path(Online Only)'], checklist['MSD Path(Online Capable)'], checklist['qVSDC Path']]:
        for cell in checklist_sheet['B']:
            condition = cell.value
            # if format_parenthesis(condition) == False:
            #     raw_input('Please modify the parenthesis in this row correctly:' + str(cell.row))
            if condition is not None:
                condition = condition.lower()
                condition = multi_replace(condition, '\n', ' ', '  ')
                condition = multi_replace(condition, '[', '', ']', '(', ')')
                condition = condition.replace('disbaled', 'disabled')
                condition = multi_replace(condition, ' disabled or not supported', ',', ' not supported', ' supported')
                condition = condition.strip().strip(',').split(',')
                # catch all configuration items
                for str in condition:
                    key = str.strip()
                    if key not in list:
                        list.append(key)
                        print key,cell
    checklist.close()